import pandas as pd
import datetime
import openpyxl
from django.db.models import Q, Sum, Count
from django.db.models.functions import ExtractMonth, ExtractYear
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from datetime import datetime, date
from django.db.models.functions import TruncMonth
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import get_user_model
from .models import MetaMensual
from apps.clientes.models import Institucion
from apps.productos.models import Producto
from .models import Pedido, DetallePedido, MetaMensual
import calendar
from django.shortcuts import get_object_or_404
from django.db import transaction # Permite deshacer cambios si hay un error
from django.http import HttpResponse




# ==========================================================
# VARIABLES GLOBALES
# ==========================================================
NOMBRES_MESES = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio', 
    7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
}

MAPEO_MESES_EXCEL = {
    'ene': 1, 'enero': 1, 'feb': 2, 'febrero': 2, 'mar': 3, 'marzo': 3,
    'abr': 4, 'abril': 4, 'may': 5, 'mayo': 5, 'jun': 6, 'junio': 6,
    'jul': 7, 'julio': 7, 'ago': 8, 'agosto': 8, 'sep': 9, 'septiembre': 9,
    'oct': 10, 'octubre': 10, 'nov': 11, 'noviembre': 11, 'dic': 12, 'diciembre': 12
}

Usuario = get_user_model()

@login_required(login_url='usuarios:login')
def metas_view(request):
    hoy_real = timezone.now().date()
    mes_sel = int(request.GET.get('mes', hoy_real.month))
    anio_sel = int(request.GET.get('anio', hoy_real.year))

    # =========================================================
    # 1. EL CEREBRO DE LA JERARQUÍA (¿A quién puedo ver?)
    # =========================================================
    if request.user.is_superuser:
        # El Dueño del sistema ve las ventas de toda la base de datos
        vendedores_permitidos = Usuario.objects.all()
    elif request.user.es_gerente:
        # El Gerente ve sus propias ventas (si tiene) + las de su equipo
        vendedores_permitidos = Usuario.objects.filter(Q(id=request.user.id) | Q(jefe=request.user))
    else:
        # El Vendedor normal solo se ve a sí mismo
        vendedores_permitidos = Usuario.objects.filter(id=request.user.id)

    # =========================================================
    # 2. CÁLCULOS APLICANDO EL FILTRO JERÁRQUICO
    # =========================================================
    
    # Sumamos las metas de todos los vendedores permitidos
    meta_total_mes = MetaMensual.objects.filter(
        representante__in=vendedores_permitidos, 
        mes=mes_sel, 
        anio=anio_sel
    ).aggregate(total=Sum('monto_meta'))['total'] or 0

    # Filtramos los pedidos solo de los vendedores permitidos
    ventas_base = Pedido.objects.filter(
        representante__in=vendedores_permitidos,
        fecha_creacion__year=anio_sel,
        fecha_creacion__month=mes_sel,
        estado__in=['Ingresado', 'Facturado']
    )

    ventas_actuales = ventas_base.aggregate(total=Sum('total_final'))['total'] or 0

    # Agrupamos por institución para la tabla inferior
    ventas_por_institucion = ventas_base.values(
        'institucion__id', 'institucion__nombre', 'institucion__rut'
    ).annotate(
        total_vendido=Sum('total_final'),
        cantidad_ordenes=Count('id')
    ).order_by('-total_vendido')

    # 3. MATEMÁTICAS (Porcentajes, brechas, ritmos)
    porcentaje = int((ventas_actuales / meta_total_mes) * 100) if meta_total_mes > 0 else 0
    brecha = max(0, meta_total_mes - ventas_actuales)
    
    dias_totales = calendar.monthrange(anio_sel, mes_sel)[1]
    dias_restantes = max(0, dias_totales - hoy_real.day) if mes_sel == hoy_real.month else 0
    ritmo = int(brecha / dias_restantes) if dias_restantes > 0 else 0
    
    bono_maximo = 500000 
    bono_estimado = int((porcentaje / 100) * bono_maximo) if porcentaje <= 100 else bono_maximo
    
    contexto = {
        'mes_sel': mes_sel,
        'anio_sel': anio_sel,
        'mes_nombre': NOMBRES_MESES.get(mes_sel),
        'nombres_meses': NOMBRES_MESES,
        'ventas_actuales': ventas_actuales,
        'meta_mensual': meta_total_mes,
        'porcentaje_avance': min(porcentaje, 100),
        'brecha': brecha,
        'ritmo_requerido': ritmo,
        'bono_estimado': bono_estimado,
        'dias_restantes': dias_restantes,
        'ventas_por_institucion': ventas_por_institucion, 
    }
    
    return render(request, 'ventas/metas.html', contexto)

#FUNCION EXCEL

# ---------------------------------------------------------
# 1. EL "GUARDIA DE SEGURIDAD"
# ---------------------------------------------------------
def es_gerente_check(user):
    # Preguntamos si el usuario existe y si su propiedad 'es_gerente' es Verdadera
    return user.is_authenticated and user.es_gerente


#FUNCION PEDIDOS

@login_required(login_url='usuarios:login')
def nueva_orden_view(request):
    # ==========================================
    # 1. LÓGICA DE PERMISOS (Jerarquía)
    # ==========================================
    if request.user.is_superuser:
        vendedores_permitidos = Usuario.objects.all()
    elif getattr(request.user, 'es_gerente', False) or request.user.groups.filter(name='Gerente').exists():
        vendedores_permitidos = Usuario.objects.filter(Q(id=request.user.id) | Q(jefe=request.user))
    else:
        vendedores_permitidos = Usuario.objects.filter(id=request.user.id)

    # ==========================================
    # 2. PROCESAMIENTO DEL FORMULARIO (POST)
    # ==========================================
    if request.method == 'POST':
        institucion_id = request.POST.get('institucion')
        notas = request.POST.get('notas_internas', '')
        contacto_id = request.POST.get('contacto')
        
        if not contacto_id or str(contacto_id).strip() == "":
            contacto_id = None
        
        productos_ids = request.POST.getlist('producto_id[]')
        cantidades = request.POST.getlist('cantidad[]')
        
        if not institucion_id or not productos_ids:
            messages.error(request, "Error: Faltan datos para crear la orden.")
            return redirect('ventas:nueva_orden')

        # EL TRUCO MAGICO: transaction.atomic() asegura que si algo falla, 
        # NADA se guarda, protegiendo tu base de datos y el stock.
        try:
            with transaction.atomic():
                # A. Guardar Cabecera
                pedido = Pedido.objects.create(
                    representante=request.user,
                    institucion_id=institucion_id,
                    contacto_id=contacto_id,
                    notas_internas=notas,
                    estado='Ingresado'
                )
                
                # B. Guardar Detalles y Validar Stock
                for prod_id, cant in zip(productos_ids, cantidades):
                    # select_for_update() bloquea la fila por milisegundos para evitar 
                    # que dos vendedores compren el último producto al mismo tiempo
                    producto = Producto.objects.select_for_update().get(id=prod_id)
                    cantidad_int = int(cant)
                    
                    # VALIDACIÓN DE STOCK
                    if cantidad_int > producto.stock:
                        raise ValueError(f"No hay stock suficiente de '{producto.nombre}'. Solicitado: {cantidad_int}, Disponible: {producto.stock}.")
                        
                    DetallePedido.objects.create(
                        pedido=pedido,
                        producto=producto,
                        cantidad=cantidad_int,
                        precio_unitario_historico=producto.precio_lista,
                        descuento_aplicado=0
                    )
                    
                    # C. Descontar Stock
                    producto.stock -= cantidad_int
                    producto.save()
                    
            # Si todo salió bien, enviamos el mensaje y redirigimos
            messages.success(request, f"¡Orden #{pedido.id} guardada exitosamente!")
            return redirect('ventas:detalle_orden', pedido_id=pedido.id)
            
        except ValueError as e:
            # Capturamos el error de stock y le avisamos al vendedor
            messages.warning(request, str(e))
            return redirect('ventas:nueva_orden')
        except Exception as e:
            messages.error(request, f"Ocurrió un error al procesar la orden: {str(e)}")
            return redirect('ventas:nueva_orden')

    # ==========================================
    # 3. CARGA DE LA VISTA (GET)
    # ==========================================
    # CORRECCIÓN CLAVE: Mostrar las clínicas del equipo, sin importar si tienen visita agendada o no.
    instituciones = Institucion.objects.filter(representante__in=vendedores_permitidos).distinct().order_by('nombre')
    
    # Solo mostramos productos con stock para no ilusionar al cliente
    productos = Producto.objects.filter(stock__gt=0).order_by('nombre')
    
    context = {
        'instituciones': instituciones,
        'productos': productos
    }
    return render(request, 'ventas/nueva_orden.html', context)


#FUNCION detalle VENTAS
# FUNCION PDF VENTAS
@login_required(login_url='usuarios:login')
def detalle_orden_view(request, pedido_id):
    # =========================================================
    # 1. BÚSQUEDA DEL PEDIDO CON JERARQUÍA
    # =========================================================
    if request.user.is_superuser:
        # El Admin puede ver TODOS los pedidos (ideal para los del Excel)
        pedido = get_object_or_404(Pedido, id=pedido_id)
        
    elif getattr(request.user, 'es_gerente', False) or request.user.groups.filter(name='Gerente').exists():
        # El Gerente puede ver los pedidos de todo su equipo
        vendedores_permitidos = Usuario.objects.filter(Q(id=request.user.id) | Q(jefe=request.user))
        pedido = get_object_or_404(Pedido, id=pedido_id, representante__in=vendedores_permitidos)
        
    else:
        # El Vendedor normal solo puede ver sus propios pedidos
        pedido = get_object_or_404(Pedido, id=pedido_id, representante=request.user)
    
    # =========================================================
    # 2. CARGA DE DETALLES Y CÁLCULOS
    # =========================================================
    # Buscamos los productos dentro de ese pedido (Mantenemos tu lógica)
    detalles = DetallePedido.objects.filter(pedido=pedido)
    
    # Si los totales están en 0, los calculamos (por si acaso)
    if getattr(pedido, 'total_neto', 0) == 0:
        pedido.calcular_totales()

    context = {
        'pedido': pedido,
        'detalles': detalles,
    }
    return render(request, 'ventas/detalle_orden.html', context)

# =========================================================
# PANEL ANALÍTICO (SOLO ADMIN / GERENCIA)
# =========================================================
@user_passes_test(es_gerente_check, login_url='visitas:dashboard')
def analitica_view(request):
    hoy = timezone.now()
    anio_actual = hoy.year
    mes_actual = request.GET.get('mes') # Capturamos el mes si existe
    
    # --- CAPTURAR EL AÑO DE LA URL ---
    anio_str = request.GET.get('anio', str(hoy.year))
    try:
        anio_actual = int(anio_str)
    except ValueError:
        anio_actual = hoy.year

    anios_disponibles = [hoy.year, hoy.year - 1, hoy.year - 2, hoy.year - 3]

    # --- BASE DE DATOS FILTRADA POR AÑO ---
    pedidos_base = Pedido.objects.filter(fecha_creacion__year=anio_actual, estado__in=['Ingresado', 'Facturado'])
    
    # Si el usuario selecciona un mes, filtramos el ranking por ese mes
    if mes_actual and mes_actual != "0": # "0" será "Todo el año"
        pedidos_base = pedidos_base.filter(fecha_creacion__month=int(mes_actual))

    # Ranking de Representantes (solo cambia la data según el filtro anterior)
    ventas_reps = pedidos_base.values('representante__first_name', 'representante__last_name')\
        .annotate(total=Sum('total_final'))\
        .order_by('-total')

    rep_labels = [f"{v['representante__first_name']} {v['representante__last_name']}" for v in ventas_reps]
    rep_data = [float(v['total']) for v in ventas_reps]

    # --- KPIs PRINCIPALES ---
    total_ingresos_anio = pedidos_base.aggregate(total=Sum('total_neto'))['total'] or 0
    total_pedidos_anio = pedidos_base.count()

    # --- GRÁFICOS 1 y 2: Ventas por Mes (Dinero y Cantidad) ---
    ventas_por_mes = pedidos_base.annotate(mes=TruncMonth('fecha_creacion')).values('mes').annotate(
        total_dinero=Sum('total_neto'),
        total_pedidos=Count('id')
    ).order_by('mes')

    meses_labels = [v['mes'].strftime('%b').capitalize() for v in ventas_por_mes]
    meses_dinero = [int(v['total_dinero']) for v in ventas_por_mes]
    meses_pedidos = [int(v['total_pedidos']) for v in ventas_por_mes]

    # --- GRÁFICO 3: Top 5 Instituciones ---
    top_instituciones = pedidos_base.values('institucion__nombre').annotate(
        total=Sum('total_neto')
    ).order_by('-total')[:5]

    inst_labels = [inst['institucion__nombre'] for inst in top_instituciones]
    inst_data = [int(inst['total']) for inst in top_instituciones]

    # --- GRÁFICO 4: Rendimiento por Representante ---
    top_reps = pedidos_base.values('representante__first_name', 'representante__last_name').annotate(
        total=Sum('total_neto')
    ).order_by('-total')

    rep_labels = [f"{rep['representante__first_name']} {rep['representante__last_name']}" for rep in top_reps]
    rep_data = [int(rep['total']) for rep in top_reps]

    # --- GRÁFICO 5: Comparativa Histórica Anual (Todos los años) ---
    ventas_historicas = Pedido.objects.filter(estado__in=['Ingresado', 'Facturado']).values('fecha_creacion__year').annotate(
        total=Sum('total_neto')
    ).order_by('fecha_creacion__year')

    hist_labels = [str(v['fecha_creacion__year']) for v in ventas_historicas]
    hist_data = [int(v['total']) for v in ventas_historicas]

    # --- GRÁFICO 6: Top 5 Productos Más Vendidos ---
    # Nota: Asegúrate de que 'detalles__producto__nombre' y 'detalles__cantidad' coincidan con tus modelos
    try:
        top_productos = pedidos_base.values('detalles__producto__nombre').annotate(
            total_vendido=Sum('detalles__cantidad')
        ).order_by('-total_vendido')[:5]

        prod_labels = [p['detalles__producto__nombre'] for p in top_productos if p.get('detalles__producto__nombre')]
        prod_data = [int(p['total_vendido']) for p in top_productos if p.get('detalles__producto__nombre')]
    except Exception as e:
        # En caso de que la relación de la base de datos sea distinta, evitamos que la vista se caiga
        prod_labels = []
        prod_data = []

        # ==========================================
    # LÓGICA DEL SEMÁFORO DE GESTIÓN (Mes Actual)
    # ==========================================
    hoy = date.today()
    vendedores_en_riesgo = []

    # Calculamos qué porcentaje del mes ya pasó (ej: día 20 de 30 = 66% del mes)
    dias_en_el_mes = calendar.monthrange(hoy.year, hoy.month)[1]
    porcentaje_mes_transcurrido = (hoy.day / dias_en_el_mes) * 100

    # Solo activamos la alerta si ya pasó al menos la mitad del mes (50%)
    if porcentaje_mes_transcurrido > 50:
        metas_mes_actual = MetaMensual.objects.filter(mes=hoy.month, anio=hoy.year)
        
        for meta in metas_mes_actual:
            venta_real = Pedido.objects.filter(
                representante=meta.representante,
                fecha_creacion__month=hoy.month,
                fecha_creacion__year=hoy.year,
                estado='Facturado'
            ).aggregate(total=Sum('total_final'))['total'] or 0
            
            porcentaje_cumplimiento = (venta_real / meta.monto_meta * 100) if meta.monto_meta > 0 else 0
            
            # Si lleva menos del 40% y ya estamos a fin de mes, ¡ALERTA ROJA!
            if porcentaje_cumplimiento < 40:
                vendedores_en_riesgo.append({
                    'nombre': meta.representante.get_full_name() or meta.representante.username,
                    'cumplimiento': int(porcentaje_cumplimiento),
                    'venta': venta_real,
                    'meta': meta.monto_meta
                })

    # --- CONTEXTO ---
    contexto = {
        'vendedores_en_riesgo': vendedores_en_riesgo,
        'anio_actual': anio_actual,
        'total_ingresos_anio': total_ingresos_anio,
        'total_pedidos_anio': total_pedidos_anio,
        'anios_disponibles': anios_disponibles,
        'rep_labels': rep_labels,
        'rep_data': rep_data,
        'mes_actual': int(mes_actual) if mes_actual else 0,
        
        # Pasamos las listas crudas de Python, Django se encarga del JSON en el template
        'meses_labels': meses_labels,
        'meses_data': meses_dinero,
        'meses_pedidos': meses_pedidos,
        'inst_labels': inst_labels,
        'inst_data': inst_data,
        'rep_labels': rep_labels,
        'rep_data': rep_data,
        'hist_labels': hist_labels,
        'hist_data': hist_data,
        'prod_labels': prod_labels,
        'prod_data': prod_data,
    }
    return render(request, 'ventas/analitica.html', contexto)


#CARGAR DATOS DE EXCEL VISTA MAESTRA

User = get_user_model()
@login_required(login_url='usuarios:login')
def cargar_datos_excel(request):
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        excel_file = request.FILES['archivo_excel']
        try:
            xls = pd.ExcelFile(excel_file)
            hojas = [str(h).strip().lower() for h in xls.sheet_names] 
            
            inst_creadas = 0
            metas_agregadas = 0
            ventas_registradas = 0

            with transaction.atomic():
                # --- 1. HOJA: ASIGNACION ---
                if 'asignacion' in hojas:
                    sheet_name = next(h for h in xls.sheet_names if str(h).strip().lower() == 'asignacion')
                    df_asig = pd.read_excel(xls, sheet_name, dtype=str).fillna('')
                    df_asig.columns = [str(c).strip().lower() for c in df_asig.columns]
                    
                    for _, row in df_asig.iterrows():
                        rep = str(row.get('representante', '')).strip()
                        rut = str(row.get('rut-3', '')).strip()
                        razon = str(row.get('razon social', '')).strip()

                        if not rep or not rut: 
                            continue
                            
                        vendedor, _ = User.objects.get_or_create(username=rep, defaults={'first_name': f"Vendedor {rep}", 'is_active': True})
                        if not vendedor.password or vendedor.password == "":
                            vendedor.set_password("FarmaView.2026")
                            vendedor.save()
                            
                        if not razon:
                            razon = f"Cliente {rut}"
                            
                        Institucion.objects.update_or_create(rut=rut, defaults={'nombre': razon, 'representante': vendedor})
                        inst_creadas += 1

                # --- 2. HOJA: META CLIENTE ---
                if 'meta cliente' in hojas:
                    sheet_name = next(h for h in xls.sheet_names if str(h).strip().lower() == 'meta cliente')
                    df_metas = pd.read_excel(xls, sheet_name, dtype=str).fillna('')
                    df_metas.columns = [str(c).strip().lower() for c in df_metas.columns]
                    
                    for _, row in df_metas.iterrows():
                        rep = str(row.get('representante', '')).strip()
                        mes_txt = str(row.get('mes', '')).strip().lower()
                        
                        if not rep or not mes_txt: 
                            continue
                            
                        vendedor = User.objects.filter(username=rep).first()
                        mes_n = MAPEO_MESES_EXCEL.get(mes_txt) 
                        
                        if vendedor and mes_n:
                            try:
                                anio_val = row.get('año', '2026')
                                anio_int = int(float(anio_val)) if anio_val else 2026
                                
                                meta_val = row.get('meta', '0')
                                meta_int = int(float(meta_val)) if meta_val else 0
                                
                                meta_obj, _ = MetaMensual.objects.get_or_create(representante=vendedor, mes=mes_n, anio=anio_int, defaults={'monto_meta': 0})
                                meta_obj.monto_meta += meta_int
                                meta_obj.save()
                                metas_agregadas += 1
                            except ValueError:
                                pass

                # --- 3. HOJA: VENTA MES ---
                if 'venta mes' in hojas:
                    sheet_name = next(h for h in xls.sheet_names if str(h).strip().lower() == 'venta mes')
                    df_ventas = pd.read_excel(xls, sheet_name, dtype=str).fillna('')
                    df_ventas.columns = [str(c).strip().lower() for c in df_ventas.columns]
                    
                    for _, row in df_ventas.iterrows():
                        rut = str(row.get('rut-3', '')).strip()
                        prod_nombre = str(row.get('producto', '')).strip()
                        
                        if not rut or not prod_nombre: 
                            continue
                            
                        inst = Institucion.objects.filter(rut=rut).first()
                        
                        if inst and inst.representante:
                            precio = 0
                            unidades = 0
                            subtotal = 0
                            anio_int = timezone.now().year
                            mes_n = 1
                            
                            try:
                                pre_val = row.get('precio', '0')
                                if pre_val and pre_val.strip() != '':
                                    precio = int(float(pre_val))
                            except Exception:
                                pass
                                
                            prod, _ = Producto.objects.get_or_create(
                                nombre=prod_nombre, 
                                defaults={'stock': 1000, 'precio_lista': precio}
                            )
                            
                            try:
                                mes_txt = str(row.get('mes', '')).strip().lower()
                                mes_n = MAPEO_MESES_EXCEL.get(mes_txt, 1) 
                                
                                anio_val = row.get('año', '')
                                if anio_val and anio_val.strip() != '':
                                    anio_int = int(float(anio_val))
                                    
                                unid_val = row.get('unidades', '0')
                                if unid_val and unid_val.strip() != '':
                                    unidades = int(float(unid_val))
                                    
                                sub_val = row.get('valores', '0')
                                if sub_val and sub_val.strip() != '':
                                    subtotal = int(float(sub_val))
                            except Exception:
                                pass
                                
                            # Creamos SIEMPRE un pedido nuevo por cada fila del Excel
                            pedido = Pedido.objects.create(
                                institucion=inst, 
                                representante=inst.representante, 
                                estado='Facturado'
                            )
                            
                            # Ajustamos la fecha según el Excel
                            fecha_falsa = timezone.make_aware(datetime(anio_int, mes_n, 15))
                            Pedido.objects.filter(id=pedido.id).update(fecha_creacion=fecha_falsa)
                            pedido.refresh_from_db()

                            DetallePedido.objects.create(
                                pedido=pedido, producto=prod, cantidad=unidades, precio_unitario_historico=precio, subtotal=subtotal
                            )
                            pedido.calcular_totales()
                            ventas_registradas += 1

            msg_final = f"✅ Éxito total. Se cargaron: {inst_creadas} Instituciones, {metas_agregadas} registros de metas, y {ventas_registradas} líneas de venta."
            messages.success(request, msg_final)
            return redirect('ventas:cargar_metas') # Ajusta el nombre de la url si es diferente en tu proyecto
            
        except Exception as e:
            messages.error(request, f"Error en el código o archivo: {e}")
            
    return render(request, 'ventas/cargar_metas.html')

#FUNCION DETALLE DE INSTITUCIONES

@login_required(login_url='usuarios:login')
def detalle_institucion_view(request, institucion_id, mes, anio):
    institucion = get_object_or_404(Institucion, id=institucion_id)
    
    # 1. Buscamos los PEDIDOS de esa institución en ese mes
    pedidos = Pedido.objects.filter(
        institucion=institucion,
        fecha_creacion__year=anio,
        fecha_creacion__month=mes,
        estado__in=['Ingresado', 'Facturado']
    ).order_by('-id')

    # 2. LÓGICA DE PERMISOS (El gerente ve todo, el vendedor solo lo suyo)
    if not (request.user.is_superuser or request.user.groups.filter(name='Gerente').exists()):
        pedidos = pedidos.filter(representante=request.user)
    
    # 3. Calculamos el total sumando el total_final de todas esas órdenes
    total_mes = pedidos.aggregate(total=Sum('total_final'))['total'] or 0

    contexto = {
        'institucion': institucion,
        'pedidos': pedidos, # <-- ¡CLAVE! Ahora enviamos 'pedidos' al HTML
        'mes_nombre': NOMBRES_MESES.get(mes), 
        'anio': anio,
        'total_mes': total_mes,
    }
    return render(request, 'ventas/detalle_institucion.html', contexto)



#FUNCION DESCARGAR EXCEL DE VENTAS
def exportar_pedidos_excel(request):
    """
    Exporta el detalle de todos los pedidos (Carrito) a Excel.
    Cada fila representa un producto dentro de un pedido.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Detallado de Pedidos"

    # --- ESTILOS ---
    blue_fill = PatternFill(start_color="1784C7", end_color="1784C7", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # --- CABECERAS ---
    headers = [
        'ID Pedido', 'Fecha', 'Estado', 'Representante', 
        'Institución', 'Contacto', 'Producto', 'Cant.', 
        'Precio Unit. (Hist)', 'Desc. %', 'Subtotal Linea', 'TOTAL PEDIDO'
    ]
    
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # --- DATOS ---
    # Usamos select_related y prefetch_related para que sea rápido (Nivel Senior)
    pedidos = Pedido.objects.select_related('representante', 'institucion', 'contacto')\
                            .prefetch_related('detalles__producto').all()

    for pedido in pedidos:
        detalles = pedido.detalles.all()
        for i, detalle in enumerate(detalles):
            # Solo mostramos el "Total Pedido" en la primera fila del grupo para que sea legible
            total_mostrar = pedido.total_final if i == 0 else ""
            
            row = [
                pedido.id,
                pedido.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
                pedido.estado,
                pedido.representante.get_full_name() or pedido.representante.username,
                pedido.institucion.nombre,
                pedido.contacto.nombre if pedido.contacto else "No especificado",
                detalle.producto.nombre,
                detalle.cantidad,
                detalle.precio_unitario_historico,
                f"{detalle.descuento_aplicado}%",
                detalle.subtotal,
                total_mostrar
            ]
            ws.append(row)

    # --- AJUSTE DE COLUMNAS ---
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    # --- RESPUESTA ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Pedidos_FarmaView_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

#EXPORTAR METAS


def exportar_metas_excel(request):
    wb = Workbook()
    
    # === HOJA 1: RESUMEN DE CUMPLIMIENTO ===
    ws1 = wb.active
    ws1.title = "Cumplimiento de Metas"
    
    headers1 = ['Mes/Año', 'Representante', 'Meta Asignada ($)', 'Venta Real ($)', 'Faltante ($)', '%', 'Estado']
    ws1.append(headers1)
    
    # Estilo Naranja FarmaView
    orange_fill = PatternFill(start_color="F37A20", end_color="F37A20", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws1[1]:
        cell.fill = orange_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    # Datos Hoja 1
    metas = MetaMensual.objects.select_related('representante').all().order_by('anio', 'mes')
    for meta in metas:
        venta_real = Pedido.objects.filter(
            representante=meta.representante,
            fecha_creacion__month=meta.mes,
            fecha_creacion__year=meta.anio,
            estado='Facturado'
        ).aggregate(total=Sum('total_final'))['total'] or 0

        faltante = max(0, meta.monto_meta - venta_real)
        porcentaje = (venta_real / meta.monto_meta * 100) if meta.monto_meta > 0 else 0
        
        ws1.append([
            f"{meta.mes}/{meta.anio}",
            meta.representante.get_full_name() or meta.representante.username,
            meta.monto_meta,
            venta_real,
            faltante,
            f"{int(porcentaje)}%",
            "CUMPLIDA" if porcentaje >= 100 else "EN PROCESO"
        ])

    # Formato moneda Hoja 1
    for row in ws1.iter_rows(min_row=2, max_col=5, max_row=ws1.max_row):
        for cell in row[2:5]:
            cell.number_format = '"$"#,##0'


   # === HOJA 2: DESGLOSE POR INSTITUCIÓN (MÉTODO RECOMENDADO) ===
    ws2 = wb.create_sheet(title="Detalle por Institución")
    
    # Agregamos la columna 'Estado' al final
    headers2 = ['RUT Institución', 'Nombre Institución', 'Representante', 'Mes', 'Año', 'Monto', 'Estado']
    ws2.append(headers2)
    
    blue_fill = PatternFill(start_color="1784C7", end_color="1784C7", fill_type="solid")
    for cell in ws2[1]:
        cell.fill = blue_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Modificamos el filtro para traer Facturados y Pendientes
    ventas_detalle = Pedido.objects.filter(estado__in=['Facturado', 'Pendiente']) \
        .annotate(mes_p=ExtractMonth('fecha_creacion'), anio_p=ExtractYear('fecha_creacion')) \
        .values('institucion__rut', 'institucion__nombre', 'representante__first_name', 
                'representante__last_name', 'mes_p', 'anio_p', 'estado') \
        .annotate(total=Sum('total_final')) \
        .order_by('anio_p', 'mes_p', 'estado', 'institucion__nombre')

    meses_es = {1:'ene', 2:'feb', 3:'mar', 4:'abr', 5:'may', 6:'jun', 
                7:'jul', 8:'ago', 9:'sep', 10:'oct', 11:'nov', 12:'dic'}

    for v in ventas_detalle:
        ws2.append([
            v['institucion__rut'],
            v['institucion__nombre'],
            f"{v['representante__first_name']} {v['representante__last_name']}",
            meses_es.get(v['mes_p']),
            v['anio_p'],
            v['total'],
            v['estado'].upper() # Para que resalte: FACTURADO o PENDIENTE
        ])

    # Formato moneda en la columna F (Monto)
    for row in ws2.iter_rows(min_row=2, max_col=6, max_row=ws2.max_row):
        row[5].number_format = '"$"#,##0'

    # Ajustar anchos de columnas en ambas hojas
    for sheet in [ws1, ws2]:
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 22

    # Generar Respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Reporte_Metas_y_Detalle_{timezone.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response